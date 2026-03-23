import pdfplumber
import re
import pandas as pd
import os

# 📂 Ruta de la carpeta
carpeta = r"C:\Users\LENOVO\Desktop\FACTURAS ROSADO\MARZO"

facturas = []
filas_faltantes = []

# 🔥 FUNCIÓN: formatear factura (001003-123456 → 001-003-123456)
def formatear_factura(valor):
    try:
        match = re.match(r'(\d{3})(\d{3})-(\d+)', valor)
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        return valor
    except:
        return valor


# 🔥 FUNCIÓN 1: extraer factura desde nombre del archivo
def extraer_factura_desde_nombre(nombre_archivo):
    try:
        match = re.search(r'001003\s*-\s*\d+', nombre_archivo)
        if match:
            valor = match.group()
            valor = re.sub(r'\s*-\s*', '-', valor)
            valor = formatear_factura(valor)
            return valor
        return "NO_ENCONTRADA"
    except:
        return "NO_ENCONTRADA"


# 🔥 FUNCIÓN 2: extraer factura desde PDF (respaldo)
def extraer_factura_por_posicion(pdf):
    try:
        pagina = pdf.pages[0]

        ancho = pagina.width
        alto = pagina.height

        # 📐 Zona donde está la factura
        bbox = (
            ancho * 0.55,
            alto * 0.10,
            ancho * 0.95,
            alto * 0.30
        )

        region = pagina.within_bbox(bbox)
        texto = region.extract_text()

        if texto:
            texto = texto.replace('\n', ' ').replace('\r', ' ')
            print("📍 TEXTO ZONA FACTURA:\n", texto)

            match = re.search(r'001003\s*-\s*\d+', texto)

            if match:
                valor = match.group()
                valor = re.sub(r'\s*-\s*', '-', valor)
                valor = formatear_factura(valor)
                return valor

        return "NO_ENCONTRADA"

    except Exception as e:
        print(f"⚠️ Error extrayendo factura: {e}")
        return "NO_ENCONTRADA"


# 🔄 Recorrer PDFs
for archivo in os.listdir(carpeta):
    if archivo.lower().endswith(".pdf"):
        ruta_pdf = os.path.join(carpeta, archivo)
        print(f"\n📄 Procesando: {archivo}")

        try:
            with pdfplumber.open(ruta_pdf) as pdf:

                # 🔥 1. Intentar desde nombre
                factura_encontrada = extraer_factura_desde_nombre(archivo)

                # 🔁 2. Si falla, usar PDF
                if factura_encontrada == "NO_ENCONTRADA":
                    factura_encontrada = extraer_factura_por_posicion(pdf)

                print(f"✅ Factura detectada: {factura_encontrada}")

                facturas.append({
                    "archivo": archivo,
                    "factura": factura_encontrada
                })

                # 🔍 EXTRAER TABLAS
                for pagina in pdf.pages:
                    tablas = pagina.extract_tables()

                    for tabla in tablas:
                        if tabla and len(tabla) > 1:
                            try:
                                df = pd.DataFrame(tabla[1:], columns=tabla[0])

                                # Normalizar columnas
                                df.columns = [str(col).strip().lower() for col in df.columns]

                                # Buscar columna "faltante"
                                col_faltante = None
                                for col in df.columns:
                                    if "faltante" in col:
                                        col_faltante = col
                                        break

                                if col_faltante:
                                    df[col_faltante] = pd.to_numeric(df[col_faltante], errors="coerce")

                                    df_filtrado = df[df[col_faltante] > 0]

                                    if not df_filtrado.empty:
                                        df_filtrado = df_filtrado.copy()

                                        df_filtrado.loc[:, "archivo"] = archivo
                                        df_filtrado.loc[:, "factura"] = factura_encontrada

                                        filas_faltantes.append(df_filtrado)

                            except Exception as e:
                                print(f"⚠️ Error en tabla: {e}")

        except Exception as e:
            print(f"❌ Error abriendo PDF: {e}")

# 📊 DataFrames finales
df_facturas = pd.DataFrame(facturas)

if filas_faltantes:
    df_faltantes = pd.concat(filas_faltantes, ignore_index=True)
else:
    df_faltantes = pd.DataFrame()

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def formatear_excel(ruta_archivo):
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # 🎨 Estilos
    alineacion = Alignment(horizontal="center", vertical="center")
    fuente_encabezado = Font(bold=True, color="FFFFFF")
    fondo_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 🔥 Encabezados
    for col in ws[1]:
        col.value = str(col.value).upper()
        col.alignment = alineacion
        col.font = fuente_encabezado
        col.fill = fondo_rojo

    # 🔥 Datos
    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = alineacion

    wb.save(ruta_archivo)


# 💾 Guardar y formatear
try:
    archivo1 = "facturas_encontradas.xlsx"
    df_facturas.to_excel(archivo1, index=False)
    formatear_excel(archivo1)
    print("\n✅ facturas_encontradas.xlsx generado y formateado")
except Exception as e:
    print(f"❌ Error guardando facturas: {e}")

try:
    if not df_faltantes.empty:
        archivo2 = "faltantes_mayores_a_0.xlsx"
        df_faltantes.to_excel(archivo2, index=False)
        formatear_excel(archivo2)
        print("✅ faltantes_mayores_a_0.xlsx generado y formateado")
    else:
        print("⚠️ No se encontraron faltantes > 0")
except Exception as e:
    print(f"❌ Error guardando faltantes: {e}")
    